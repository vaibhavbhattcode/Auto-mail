import io
import csv
import json
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import database

# Try to import FPDF with a fallback
try:
    from fpdf import FPDF
except ImportError:
    FPDF = None

def get_campaign_data(campaign_id):
    """Fetch all campaign info, statistics, and leads"""
    conn = database.get_db()
    cursor = conn.cursor()
    
    # Campaign details
    cursor.execute('SELECT * FROM campaigns WHERE id = ?', (campaign_id,))
    campaign = cursor.fetchone()
    if not campaign:
        conn.close()
        return None
    campaign = dict(campaign)
    
    # Leads list
    cursor.execute('SELECT * FROM leads WHERE campaign_id = ? ORDER BY id ASC', (campaign_id,))
    leads = [dict(row) for row in cursor.fetchall()]
    
    # Stats calculation
    total = len(leads)
    sent = sum(1 for l in leads if l['status'] == 'Sent')
    pending = sum(1 for l in leads if l['status'] == 'Pending')
    scheduled = sum(1 for l in leads if l['status'] == 'Scheduled')
    failed = sum(1 for l in leads if l['status'] == 'Failed')
    
    opens = sum(1 for l in leads if l['open_count'] > 0)
    clicks = sum(1 for l in leads if l['click_count'] > 0)
    bounces = sum(1 for l in leads if l['bounce_status'] != 'none')
    unsubscribes = sum(1 for l in leads if l['unsubscribed'] > 0)
    
    success_rate = round((sent / total * 100), 1) if total > 0 else 0
    open_rate = round((opens / sent * 100), 1) if sent > 0 else 0
    click_rate = round((clicks / sent * 100), 1) if sent > 0 else 0
    bounce_rate = round((bounces / total * 100), 1) if total > 0 else 0
    
    stats = {
        'total': total,
        'sent': sent,
        'pending': pending,
        'scheduled': scheduled,
        'failed': failed,
        'opens': opens,
        'clicks': clicks,
        'bounces': bounces,
        'unsubscribes': unsubscribes,
        'success_rate': success_rate,
        'open_rate': open_rate,
        'click_rate': click_rate,
        'bounce_rate': bounce_rate
    }
    
    conn.close()
    return {
        'campaign': campaign,
        'leads': leads,
        'stats': stats
    }

def generate_csv(campaign_id):
    """Generate CSV report for campaign"""
    data = get_campaign_data(campaign_id)
    if not data:
        return None
        
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        'Lead ID', 'Recipient Email', 'CC', 'BCC', 'Subject', 
        'Status', 'Scheduled Time', 'Sent Time', 'Bounce Status', 
        'Bounce Reason', 'Unsubscribed', 'Opens', 'Clicks', 'Error Message'
    ])
    
    for l in data['leads']:
        writer.writerow([
            l['id'], l['recipient_email'], l['cc'] or '', l['bcc'] or '', l['subject'],
            l['status'], l['scheduled_time'] or '', l['sent_time'] or '', l['bounce_status'],
            l['bounce_reason'] or '', 'Yes' if l['unsubscribed'] else 'No', 
            l['open_count'], l['click_count'], l['error_message'] or ''
        ])
        
    return output.getvalue().encode('utf-8')

def generate_excel(campaign_id):
    """Generate dynamic Excel report with multiple sheets and nice styles"""
    data = get_campaign_data(campaign_id)
    if not data:
        return None
        
    wb = openpyxl.Workbook()
    
    # Sheet 1: Campaign Overview
    ws_summary = wb.active
    ws_summary.title = "Overview"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Styling helpers
    navy_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    light_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    sub_header_font = Font(name="Arial", size=12, bold=True, color="1F2937")
    bold_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # Write Title block
    ws_summary.merge_cells("A1:D1")
    ws_summary["A1"] = f"Campaign Summary: {data['campaign']['name']}"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].fill = navy_fill
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 40
    
    # Campaign Meta
    ws_summary["A3"] = "Campaign Property"
    ws_summary["A3"].font = bold_font
    ws_summary["B3"] = "Value"
    ws_summary["B3"].font = bold_font
    
    meta_rows = [
        ("Description", data['campaign']['description'] or "N/A"),
        ("Status", data['campaign']['status'].capitalize()),
        ("Start Date / Time", data['campaign']['start_time'] or "N/A"),
        ("Created At", data['campaign']['created_at']),
        ("Sending Interval", f"{data['campaign']['sending_interval']} minutes" if data['campaign']['sending_interval'] > 0 else "Individual")
    ]
    
    for r_idx, (prop, val) in enumerate(meta_rows, start=4):
        ws_summary.cell(row=r_idx, column=1, value=prop).font = bold_font
        ws_summary.cell(row=r_idx, column=2, value=val).font = normal_font
        ws_summary.cell(row=r_idx, column=1).border = thin_border
        ws_summary.cell(row=r_idx, column=2).border = thin_border
        
    # Metrics
    ws_summary["A11"] = "Performance Metric"
    ws_summary["A11"].font = bold_font
    ws_summary["B11"] = "Count"
    ws_summary["B11"].font = bold_font
    ws_summary["C11"] = "Percentage / Rate"
    ws_summary["C11"].font = bold_font
    
    metric_rows = [
        ("Total Leads", data['stats']['total'], "100.0%"),
        ("Sent Successfully", data['stats']['sent'], f"{data['stats']['success_rate']}%"),
        ("Pending", data['stats']['pending'], f"{round((data['stats']['pending']/data['stats']['total']*100), 1) if data['stats']['total'] > 0 else 0}%"),
        ("Scheduled", data['stats']['scheduled'], f"{round((data['stats']['scheduled']/data['stats']['total']*100), 1) if data['stats']['total'] > 0 else 0}%"),
        ("Failed", data['stats']['failed'], f"{round((data['stats']['failed']/data['stats']['total']*100), 1) if data['stats']['total'] > 0 else 0}%"),
        ("Bounces (Hard/Soft)", data['stats']['bounces'], f"{data['stats']['bounce_rate']}%"),
        ("Unsubscribes", data['stats']['unsubscribes'], f"{round((data['stats']['unsubscribes']/data['stats']['total']*100), 1) if data['stats']['total'] > 0 else 0}%"),
        ("Total Opens", data['stats']['opens'], f"{data['stats']['open_rate']}% Open Rate"),
        ("Total Link Clicks", data['stats']['clicks'], f"{data['stats']['click_rate']}% Click Rate"),
    ]
    
    for r_idx, (metric, count, pct) in enumerate(metric_rows, start=12):
        ws_summary.cell(row=r_idx, column=1, value=metric).font = bold_font
        ws_summary.cell(row=r_idx, column=2, value=count).font = normal_font
        ws_summary.cell(row=r_idx, column=3, value=pct).font = normal_font
        
        # Style highlighted metrics
        if "Rate" in pct or "Success" in metric:
            ws_summary.cell(row=r_idx, column=1).fill = light_blue_fill
            ws_summary.cell(row=r_idx, column=2).fill = light_blue_fill
            ws_summary.cell(row=r_idx, column=3).fill = light_blue_fill
            
        ws_summary.cell(row=r_idx, column=1).border = thin_border
        ws_summary.cell(row=r_idx, column=2).border = thin_border
        ws_summary.cell(row=r_idx, column=3).border = thin_border

    # Adjust summary column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 30
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 20
    
    # Sheet 2: Leads Details
    ws_leads = wb.create_sheet(title="Recipient Details")
    ws_leads.views.sheetView[0].showGridLines = True
    
    headers = [
        'Lead ID', 'Recipient Email', 'CC', 'BCC', 'Subject', 
        'Status', 'Scheduled Time', 'Sent Time', 'Bounce Status', 
        'Opens', 'Clicks', 'Unsubscribed', 'Error Message'
    ]
    
    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_leads.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_leads.row_dimensions[1].height = 25
    
    # Write lead rows
    for r_idx, l in enumerate(data['leads'], start=2):
        ws_leads.cell(row=r_idx, column=1, value=l['id']).font = normal_font
        ws_leads.cell(row=r_idx, column=2, value=l['recipient_email']).font = normal_font
        ws_leads.cell(row=r_idx, column=3, value=l['cc'] or '').font = normal_font
        ws_leads.cell(row=r_idx, column=4, value=l['bcc'] or '').font = normal_font
        ws_leads.cell(row=r_idx, column=5, value=l['subject']).font = normal_font
        
        status_cell = ws_leads.cell(row=r_idx, column=6, value=l['status'])
        status_cell.font = bold_font
        if l['status'] == 'Sent':
            status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # light green
        elif l['status'] == 'Failed':
            status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # light red
        elif l['status'] == 'Scheduled':
            status_cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid") # light blue
            
        ws_leads.cell(row=r_idx, column=7, value=l['scheduled_time'] or '').font = normal_font
        ws_leads.cell(row=r_idx, column=8, value=l['sent_time'] or '').font = normal_font
        ws_leads.cell(row=r_idx, column=9, value=l['bounce_status']).font = normal_font
        ws_leads.cell(row=r_idx, column=10, value=l['open_count']).font = normal_font
        ws_leads.cell(row=r_idx, column=11, value=l['click_count']).font = normal_font
        ws_leads.cell(row=r_idx, column=12, value='Yes' if l['unsubscribed'] else 'No').font = normal_font
        ws_leads.cell(row=r_idx, column=13, value=l['error_message'] or '').font = normal_font
        
        # Border cells
        for c in range(1, 14):
            ws_leads.cell(row=r_idx, column=c).border = thin_border
            
    # Autofit column widths
    for col in ws_leads.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_leads.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
        
    # Stream workbook
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()

class PDFReport(FPDF):
    """Custom FPDF document layout with nice header and footer"""
    def header(self):
        # Draw top navbar color strip
        self.set_fill_color(31, 41, 55) # Navy
        self.rect(0, 0, 210, 8, 'F')
        
        # Brand logo / header
        self.set_y(15)
        self.set_font('Arial', 'B', 15)
        self.set_text_color(31, 41, 55)
        self.cell(0, 8, 'Bhatt Technologies Outreach Campaign Report', 0, 1, 'L')
        self.set_font('Arial', '', 9)
        self.set_text_color(107, 114, 128) # Grey
        self.cell(0, 5, 'Outreach Automation & Tracking Engine', 0, 1, 'L')
        self.ln(5)
        self.line(10, 32, 200, 32)
        
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(156, 163, 175)
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}} | Generated on {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}', 0, 0, 'C')

def generate_pdf(campaign_id):
    """Generate styled PDF campaign summary report using fpdf2"""
    data = get_campaign_data(campaign_id)
    if not data:
        return None
        
    if not FPDF:
        # Fallback to plain text masqueraded as PDF/download or text representation
        return b"Error: FPDF2 library is not installed on this server. Please export as Excel or CSV."
        
    pdf = PDFReport()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=20)
    
    # 1. Overview Section
    pdf.set_y(38)
    pdf.set_font('Arial', 'B', 12)
    pdf.set_text_color(31, 41, 55)
    pdf.cell(0, 8, 'Campaign Metadata', 0, 1, 'L')
    
    pdf.set_font('Arial', '', 10)
    pdf.set_fill_color(249, 250, 251)
    
    # Render table properties
    meta_props = [
        ('Campaign Name', data['campaign']['name']),
        ('Description', data['campaign']['description'] or 'None'),
        ('Status', data['campaign']['status'].upper()),
        ('Scheduled Start', data['campaign']['start_time'] or 'Immediate'),
        ('Sending Interval', f"{data['campaign']['sending_interval']} minutes" if data['campaign']['sending_interval'] > 0 else 'None'),
        ('Created Date', data['campaign']['created_at'])
    ]
    
    for prop, val in meta_props:
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(45, 7, f" {prop}", 1, 0, 'L', fill=True)
        pdf.set_font('Arial', '', 9)
        pdf.cell(145, 7, f" {val}", 1, 1, 'L')
        
    pdf.ln(8)
    
    # 2. Performance Metrics Grid (Card style block)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 8, 'Performance & Delivery Statistics', 0, 1, 'L')
    
    # Draw stats card grid
    stats_cols = [
        [('Total Leads', data['stats']['total']), ('Sent successfully', data['stats']['sent']), ('Pending Queue', data['stats']['pending'])],
        [('Failed Emails', data['stats']['failed']), ('Bounces Recorded', data['stats']['bounces']), ('Unsubscribes', data['stats']['unsubscribes'])],
        [('Open Rate (%)', f"{data['stats']['open_rate']}%"), ('Click Rate (%)', f"{data['stats']['click_rate']}%"), ('Success Rate (%)', f"{data['stats']['success_rate']}%")]
    ]
    
    pdf.set_font('Arial', '', 9)
    for row in stats_cols:
        for title, val in row:
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(63, 6, f" {title}", 'LTR', 0, 'C', fill=True)
        pdf.cell(0, 6, '', 0, 1) # newline
        for title, val in row:
            pdf.set_font('Arial', 'B', 14)
            pdf.set_text_color(37, 99, 235) # Blue text
            pdf.cell(63, 10, f"{val}", 'LBR', 0, 'C')
            pdf.set_text_color(31, 41, 55) # reset
        pdf.ln(13)
        
    pdf.ln(5)
    
    # 3. Deliverability Breakdown Table
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 8, 'Detailed Status Summary', 0, 1, 'L')
    
    # Table headers
    pdf.set_font('Arial', 'B', 9)
    pdf.set_text_color(255, 255, 255)
    pdf.set_fill_color(31, 41, 55)
    pdf.cell(15, 8, ' ID', 1, 0, 'C', fill=True)
    pdf.cell(60, 8, ' Recipient Email', 1, 0, 'L', fill=True)
    pdf.cell(35, 8, ' Status', 1, 0, 'C', fill=True)
    pdf.cell(20, 8, ' Opens', 1, 0, 'C', fill=True)
    pdf.cell(20, 8, ' Clicks', 1, 0, 'C', fill=True)
    pdf.cell(40, 8, ' Sent Date', 1, 1, 'C', fill=True)
    
    pdf.set_text_color(31, 41, 55)
    pdf.set_font('Arial', '', 8.5)
    
    # Max rows for preview inside PDF, if too long truncate to prevent 100 pages of PDF
    show_leads = data['leads'][:30]
    
    for idx, l in enumerate(show_leads):
        fill = (idx % 2 == 1)
        pdf.set_fill_color(249, 250, 251)
        pdf.cell(15, 7, f" {l['id']}", 1, 0, 'C', fill=fill)
        pdf.cell(60, 7, f" {l['recipient_email']}", 1, 0, 'L', fill=fill)
        
        # Color code status
        if l['status'] == 'Sent':
            pdf.set_text_color(16, 185, 129) # green
        elif l['status'] == 'Failed':
            pdf.set_text_color(239, 68, 68) # red
        else:
            pdf.set_text_color(245, 158, 11) # amber
            
        pdf.cell(35, 7, f" {l['status']}", 1, 0, 'C', fill=fill)
        pdf.set_text_color(31, 41, 55) # reset
        
        pdf.cell(20, 7, f" {l['open_count']}", 1, 0, 'C', fill=fill)
        pdf.cell(20, 7, f" {l['click_count']}", 1, 0, 'C', fill=fill)
        pdf.cell(40, 7, f" {l['sent_time'] or 'N/A'}", 1, 1, 'C', fill=fill)
        
    if len(data['leads']) > 30:
        pdf.set_font('Arial', 'I', 8.5)
        pdf.cell(0, 7, f"... and {len(data['leads']) - 30} more recipients. Full list available in Excel/CSV exports.", 0, 1, 'C')
        
    # Output to bytes
    return bytes(pdf.output())
